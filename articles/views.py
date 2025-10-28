from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.template.loader import render_to_string
from django.db import transaction
from django.utils import timezone
from .models import Categorie, Article
from .forms import CategorieForm, ArticleForm
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io
import os


# ===== VUES POUR LES CATÉGORIES =====

@login_required
def categorie_list(request):
    """Page de gestion des catégories avec formulaire d'ajout et liste"""
    from django.core.paginator import Paginator
    from django.contrib import messages
    
    # Création via formulaire inline
    if request.method == 'POST':
        form = CategorieForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Catégorie ajoutée avec succès')
            return redirect('articles:categorie_list')
        else:
            messages.error(request, 'Erreur lors de l\'ajout de la catégorie')
    else:
        # Formulaire vide pour les requêtes GET
        form = CategorieForm()
    
    # Récupération des catégories avec filtres
    categories_qs = Categorie.objects.all()
    
    # Recherche par libellé
    search_query = request.GET.get('search', '').strip()
    if search_query:
        categories_qs = categories_qs.filter(libelle__icontains=search_query)
    
    # Tri
    sort_by = request.GET.get('sort', 'libelle')
    sort_order = request.GET.get('order', 'asc')
    valid_sort_fields = ['libelle', 'date_creation', 'date_modification', 'actif']
    if sort_by not in valid_sort_fields:
        sort_by = 'libelle'
    if sort_order == 'desc':
        sort_by = f'-{sort_by}'
    categories_qs = categories_qs.order_by(sort_by)
    
    # Filtre par statut
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'active':
        categories_qs = categories_qs.filter(actif=True)
    elif status_filter == 'inactive':
        categories_qs = categories_qs.filter(actif=False)
    
    # Pagination
    page_number = request.GET.get('page', 1)
    paginator = Paginator(categories_qs, 10)
    page_obj = paginator.get_page(page_number)
    
    # Stats
    active_categories_count = categories_qs.filter(actif=True).count()
    inactive_categories_count = categories_qs.filter(actif=False).count()
    
    context = {
        'form': form,
        'categories': page_obj.object_list,
        'page_obj': page_obj,
        'paginator': paginator,
        'active_categories_count': active_categories_count,
        'inactive_categories_count': inactive_categories_count,
        'search_query': search_query,
        'sort_by': sort_by.replace('-', '') if isinstance(sort_by, str) and sort_by.startswith('-') else sort_by,
        'sort_order': sort_order,
        'status_filter': status_filter,
    }
    
    return render(request, 'articles/categorie_page.html', context)


@login_required
def categorie_create_popup(request):
    """Créer une catégorie via popup"""
    if request.method == 'POST':
        form = CategorieForm(request.POST)
        if form.is_valid():
            categorie = form.save()
            return JsonResponse({
                'success': True,
                'message': 'Catégorie créée avec succès',
                'categorie_id': categorie.id,
                'categorie_libelle': categorie.libelle
            })
        else:
            html = render_to_string('articles/categorie_form_modal.html', {
                'form': form,
                'title': 'Nouvelle catégorie'
            }, request=request)
            return JsonResponse({
                'success': False,
                'html': html
            })
    else:
        form = CategorieForm()
        html = render_to_string('articles/categorie_form_modal.html', {
            'form': form,
            'title': 'Nouvelle catégorie'
        }, request=request)
        return JsonResponse({'html': html})


@login_required
def categorie_update_popup(request, pk):
    """Modifier une catégorie via popup"""
    categorie = get_object_or_404(Categorie, pk=pk)
    
    if request.method == 'POST':
        form = CategorieForm(request.POST, instance=categorie)
        if form.is_valid():
            categorie = form.save()
            return JsonResponse({
                'success': True,
                'message': 'Catégorie modifiée avec succès',
                'categorie_id': categorie.id,
                'categorie_libelle': categorie.libelle
            })
        else:
            html = render_to_string('articles/categorie_form_modal.html', {
                'form': form,
                'title': 'Modifier la catégorie',
                'categorie': categorie
            }, request=request)
            return JsonResponse({
                'success': False,
                'html': html
            })
    else:
        form = CategorieForm(instance=categorie)
        html = render_to_string('articles/categorie_form_modal.html', {
            'form': form,
            'title': 'Modifier la catégorie',
            'categorie': categorie
        }, request=request)
        return JsonResponse({'html': html})


@login_required
def categorie_delete_popup(request, pk):
    """Supprimer une catégorie via popup"""
    categorie = get_object_or_404(Categorie, pk=pk)
    
    if request.method == 'POST':
        try:
            libelle = categorie.libelle
            categorie.delete()
            return JsonResponse({
                'success': True,
                'message': f'Catégorie "{libelle}" supprimée avec succès'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Erreur lors de la suppression: {str(e)}'
            })
    else:
        html = render_to_string('articles/categorie_confirm_delete_modal.html', {
            'categorie': categorie
        }, request=request)
        return JsonResponse({'html': html})


# ===== VUES POUR LES ARTICLES =====

@login_required
def article_list(request):
    """Liste des articles avec recherche et tri"""
    articles = Article.objects.select_related('categorie').all()
    
    # Recherche par désignation ou catégorie
    search_query = request.GET.get('search', '').strip()
    if search_query:
        articles = articles.filter(
            models.Q(designation__icontains=search_query) |
            models.Q(categorie__libelle__icontains=search_query)
        )
    
    # Tri
    sort_by = request.GET.get('sort', 'designation')
    sort_order = request.GET.get('order', 'asc')
    
    valid_sort_fields = ['designation', 'categorie__libelle', 'date_creation', 'date_modification', 'actif']
    if sort_by not in valid_sort_fields:
        sort_by = 'designation'
    
    if sort_order == 'desc':
        sort_by = f'-{sort_by}'
    
    articles = articles.order_by(sort_by)
    
    # Filtre par statut
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'active':
        articles = articles.filter(actif=True)
    elif status_filter == 'inactive':
        articles = articles.filter(actif=False)
    
    # Filtre par catégorie
    category_filter = request.GET.get('category', 'all')
    if category_filter != 'all':
        articles = articles.filter(categorie_id=category_filter)
    
    # Récupérer toutes les catégories pour le filtre
    categories = Categorie.objects.filter(actif=True).order_by('libelle')
    
    # Compter les articles actifs
    active_articles_count = articles.filter(actif=True).count()
    
    return render(request, 'articles/article_list.html', {
        'articles': articles,
        'categories': categories,
        'active_articles_count': active_articles_count,
        'search_query': search_query,
        'sort_by': sort_by.replace('-', '') if sort_by.startswith('-') else sort_by,
        'sort_order': sort_order,
        'status_filter': status_filter,
        'category_filter': category_filter,
    })


@login_required
def article_liste_print(request):
    """Vue pour l'aperçu PDF de la liste des articles"""
    from datetime import datetime
    from django.db import models
    
    articles = Article.objects.select_related('categorie').all()
    
    # Recherche par désignation ou catégorie
    search_query = request.GET.get('search', '').strip()
    if search_query:
        articles = articles.filter(
            models.Q(designation__icontains=search_query) |
            models.Q(categorie__libelle__icontains=search_query)
        )
    
    # Filtre par statut
    status_filter = request.GET.get('status', 'all')
    if status_filter == 'active':
        articles = articles.filter(actif=True)
    elif status_filter == 'inactive':
        articles = articles.filter(actif=False)
    
    # Filtre par catégorie
    category_filter = request.GET.get('category', 'all')
    if category_filter != 'all':
        articles = articles.filter(categorie_id=category_filter)
    
    # Tri par défaut
    articles = articles.order_by('categorie__libelle', 'designation')
    
    # Statistiques
    stats = {
        'en_stock': articles.filter(stock_actuel__gt=0).count(),
        'rupture_stock': articles.filter(stock_actuel=0).count(),
        'stock_faible': articles.filter(stock_actuel__lte=models.F('stock_minimum')).count(),
        'valeur_stock': sum(article.prix_unitaire * article.stock_actuel for article in articles if article.prix_unitaire and article.stock_actuel),
    }
    
    # Contexte pour le template
    context = {
        'articles': articles,
        'societe': get_societe_info(),
        'document_type': 'Liste des articles',
        'date_debut': datetime.now().date(),
        'date_fin': datetime.now().date(),
        'filtre_categorie': category_filter or 'Toutes',
        'stats': stats,
    }
    
    return render(request, 'articles/liste_articles_print.html', context)

def get_societe_info():
    """Retourne les informations de l'entreprise"""
    return {
        'nom': 'DEVDRECO SOFT',
        'adresse': '123 Avenue des Affaires',
        'ville': 'Lomé, Togo',
        'telephone': '+228 90 12 34 56',
        'email': 'contact@devdreco-soft.com',
        'site_web': 'www.devdreco-soft.com',
        'logo': None,
    }


@login_required
def article_designations_api(request):
    """Retourne la liste des désignations d'articles (JSON) pour l'autocomplétion."""
    designations = list(
        Article.objects.filter(actif=True)
        .order_by('designation')
        .values_list('designation', flat=True)
    )
    return JsonResponse({'designations': designations})

@login_required
def article_create_popup(request):
    """Créer un article via popup"""
    if request.method == 'POST':
        form = ArticleForm(request.POST)
        if form.is_valid():
            article = form.save()
            return JsonResponse({
                'success': True,
                'message': 'Article créé avec succès',
                'article_id': article.id,
                'article_designation': article.designation,
                'categorie_libelle': article.categorie.libelle
            })
        else:
            html = render_to_string('articles/article_form_modal.html', {
                'form': form,
                'title': 'Nouvel article'
            }, request=request)
            return JsonResponse({
                'success': False,
                'html': html
            })
    else:
        form = ArticleForm()
        html = render_to_string('articles/article_form_modal.html', {
            'form': form,
            'title': 'Nouvel article'
        }, request=request)
        return JsonResponse({'html': html})


@login_required
def article_update_popup(request, pk):
    """Modifier un article via popup"""
    article = get_object_or_404(Article, pk=pk)
    
    if request.method == 'POST':
        form = ArticleForm(request.POST, instance=article)
        if form.is_valid():
            article = form.save()
            return JsonResponse({
                'success': True,
                'message': 'Article modifié avec succès',
                'article_id': article.id,
                'article_designation': article.designation,
                'categorie_libelle': article.categorie.libelle
            })
        else:
            html = render_to_string('articles/article_form_modal.html', {
                'form': form,
                'title': 'Modifier l\'article',
                'article': article
            }, request=request)
            return JsonResponse({
                'success': False,
                'html': html
            })
    else:
        form = ArticleForm(instance=article)
        html = render_to_string('articles/article_form_modal.html', {
            'form': form,
            'title': 'Modifier l\'article',
            'article': article
        }, request=request)
        return JsonResponse({'html': html})


@login_required
def article_delete_popup(request, pk):
    """Supprimer un article via popup"""
    article = get_object_or_404(Article, pk=pk)
    
    if request.method == 'POST':
        try:
            designation = article.designation
            article.delete()
            return JsonResponse({
                'success': True,
                'message': f'Article "{designation}" supprimé avec succès'
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Erreur lors de la suppression: {str(e)}'
            })
    else:
        html = render_to_string('articles/article_confirm_delete_modal.html', {
            'article': article
        }, request=request)
        return JsonResponse({'html': html})


# ===== VUES D'IMPORT/EXPORT =====

@login_required
def export_categories_excel(request):
    """Exporte toutes les catégories vers un fichier Excel"""
    try:
        # Créer un nouveau workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catégories"
        
        # En-têtes
        headers = ['ID', 'Libellé', 'Actif', 'Date de création', 'Date de modification']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        categories = Categorie.objects.all().order_by('libelle')
        for row, categorie in enumerate(categories, 2):
            ws.cell(row=row, column=1, value=categorie.id)
            ws.cell(row=row, column=2, value=categorie.libelle)
            ws.cell(row=row, column=3, value="Oui" if categorie.actif else "Non")
            ws.cell(row=row, column=4, value=categorie.date_creation.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=5, value=categorie.date_modification.strftime('%d/%m/%Y %H:%M'))
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Préparer la réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="categories_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Sauvegarder dans la réponse
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Erreur lors de l\'export: {str(e)}')
        return redirect('articles:categorie_list')


@login_required
def export_articles_excel(request):
    """Exporte tous les articles vers un fichier Excel"""
    try:
        # Créer un nouveau workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Articles"
        
        # En-têtes
        headers = ['ID', 'Désignation', 'Catégorie', 'Actif', 'Date de création', 'Date de modification']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Données
        articles = Article.objects.select_related('categorie').all().order_by('designation')
        for row, article in enumerate(articles, 2):
            ws.cell(row=row, column=1, value=article.id)
            ws.cell(row=row, column=2, value=article.designation)
            ws.cell(row=row, column=3, value=article.categorie.libelle)
            ws.cell(row=row, column=4, value="Oui" if article.actif else "Non")
            ws.cell(row=row, column=5, value=article.date_creation.strftime('%d/%m/%Y %H:%M'))
            ws.cell(row=row, column=6, value=article.date_modification.strftime('%d/%m/%Y %H:%M'))
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # Préparer la réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="articles_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Sauvegarder dans la réponse
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Erreur lors de l\'export: {str(e)}')
        return redirect('articles:article_list')


@login_required
def import_categories_excel(request):
    """Importe des catégories depuis un fichier Excel"""
    if request.method == 'POST':
        try:
            uploaded_file = request.FILES.get('excel_file')
            if not uploaded_file:
                messages.error(request, 'Aucun fichier sélectionné.')
                return redirect('articles:categorie_list')
            
            # Vérifier l'extension
            if not uploaded_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Veuillez sélectionner un fichier Excel (.xlsx ou .xls).')
                return redirect('articles:categorie_list')
            
            # Lire le fichier Excel
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active
            
            # Vérifier les en-têtes
            expected_headers = ['ID', 'Libellé', 'Actif', 'Date de création', 'Date de modification']
            headers = [cell.value for cell in ws[1]]
            
            if not all(header in headers for header in ['Libellé']):
                messages.error(request, 'Format de fichier invalide. Colonnes requises: Libellé')
                return redirect('articles:categorie_list')
            
            # Traitement des données
            created_count = 0
            updated_count = 0
            errors = []
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        if not row[1]:  # Libellé vide
                            continue
                            
                        libelle = str(row[1]).strip()
                        if not libelle:
                            continue
                        
                        # Vérifier si la catégorie existe déjà
                        categorie, created = Categorie.objects.get_or_create(
                            libelle=libelle,
                            defaults={
                                'actif': True,
                                'date_creation': timezone.now(),
                                'date_modification': timezone.now()
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            # Mettre à jour si nécessaire
                            if len(row) > 2 and row[2] is not None:
                                actif = str(row[2]).lower() in ['oui', 'yes', 'true', '1', 'actif']
                                if categorie.actif != actif:
                                    categorie.actif = actif
                                    categorie.date_modification = timezone.now()
                                    categorie.save()
                                    updated_count += 1
                    
                    except Exception as e:
                        errors.append(f"Ligne {row_num}: {str(e)}")
            
            # Messages de résultat
            if created_count > 0:
                messages.success(request, f'{created_count} catégorie(s) créée(s) avec succès.')
            if updated_count > 0:
                messages.info(request, f'{updated_count} catégorie(s) mise(s) à jour.')
            if errors:
                messages.warning(request, f'Erreurs rencontrées: {len(errors)} ligne(s) ignorée(s).')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'import: {str(e)}')
    
    return redirect('articles:categorie_list')


@login_required
def import_articles_excel(request):
    """Importe des articles depuis un fichier Excel"""
    if request.method == 'POST':
        try:
            uploaded_file = request.FILES.get('excel_file')
            if not uploaded_file:
                messages.error(request, 'Aucun fichier sélectionné.')
                return redirect('articles:article_list')
            
            # Vérifier l'extension
            if not uploaded_file.name.endswith(('.xlsx', '.xls')):
                messages.error(request, 'Veuillez sélectionner un fichier Excel (.xlsx ou .xls).')
                return redirect('articles:article_list')
            
            # Lire le fichier Excel
            wb = openpyxl.load_workbook(uploaded_file)
            ws = wb.active
            
            # Vérifier les en-têtes
            headers = [cell.value for cell in ws[1]]
            
            if not all(header in headers for header in ['Désignation', 'Catégorie']):
                messages.error(request, 'Format de fichier invalide. Colonnes requises: Désignation, Catégorie')
                return redirect('articles:article_list')
            
            # Traitement des données
            created_count = 0
            updated_count = 0
            errors = []
            
            with transaction.atomic():
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    try:
                        if not row[1] or not row[2]:  # Désignation ou Catégorie vide
                            continue
                            
                        designation = str(row[1]).strip()
                        categorie_libelle = str(row[2]).strip()
                        
                        if not designation or not categorie_libelle:
                            continue
                        
                        # Récupérer ou créer la catégorie
                        try:
                            categorie = Categorie.objects.get(libelle=categorie_libelle)
                        except Categorie.DoesNotExist:
                            categorie = Categorie.objects.create(
                                libelle=categorie_libelle,
                                actif=True
                            )
                        
                        # Vérifier si l'article existe déjà
                        article, created = Article.objects.get_or_create(
                            designation=designation,
                            defaults={
                                'categorie': categorie,
                                'actif': True,
                                'date_creation': timezone.now(),
                                'date_modification': timezone.now()
                            }
                        )
                        
                        if created:
                            created_count += 1
                        else:
                            # Mettre à jour si nécessaire
                            if article.categorie != categorie:
                                article.categorie = categorie
                                article.date_modification = timezone.now()
                                article.save()
                                updated_count += 1
                    
                    except Exception as e:
                        errors.append(f"Ligne {row_num}: {str(e)}")
            
            # Messages de résultat
            if created_count > 0:
                messages.success(request, f'{created_count} article(s) créé(s) avec succès.')
            if updated_count > 0:
                messages.info(request, f'{updated_count} article(s) mis à jour.')
            if errors:
                messages.warning(request, f'Erreurs rencontrées: {len(errors)} ligne(s) ignorée(s).')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de l\'import: {str(e)}')
    
    return redirect('articles:article_list')


@login_required
def download_template_categories(request):
    """Télécharge un modèle Excel pour l'import de catégories"""
    try:
        # Créer un nouveau workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modèle Catégories"
        
        # En-têtes
        headers = ['Libellé', 'Actif']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Exemples
        examples = [
            ['Électronique', 'Oui'],
            ['Vêtements', 'Oui'],
            ['Alimentaire', 'Non'],
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        # Préparer la réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="modele_categories.xlsx"'
        
        # Sauvegarder dans la réponse
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Erreur lors de la génération du modèle: {str(e)}')
        return redirect('articles:categorie_list')


@login_required
def download_template_articles(request):
    """Télécharge un modèle Excel pour l'import d'articles"""
    try:
        # Créer un nouveau workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Modèle Articles"
        
        # En-têtes
        headers = ['Désignation', 'Catégorie', 'Actif']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Exemples
        examples = [
            ['Ordinateur portable', 'Électronique', 'Oui'],
            ['T-shirt', 'Vêtements', 'Oui'],
            ['Pomme', 'Alimentaire', 'Non'],
        ]
        
        for row, example in enumerate(examples, 2):
            for col, value in enumerate(example, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Ajuster la largeur des colonnes
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
        
        # Préparer la réponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="modele_articles.xlsx"'
        
        # Sauvegarder dans la réponse
        wb.save(response)
        return response
        
    except Exception as e:
        messages.error(request, f'Erreur lors de la génération du modèle: {str(e)}')
        return redirect('articles:article_list')